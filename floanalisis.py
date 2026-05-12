import streamlit as st
import pandas as pd
import re
import numpy as np
from openpyxl import load_workbook
from datetime import datetime, timedelta, date
import warnings
warnings.filterwarnings('ignore')

# ==========================================
# 🎨 CONFIGURACIÓN DE PÁGINA
# ==========================================
st.set_page_config(page_title="Análisis Inteligente FL Tarde", page_icon="🍑", layout="wide")
st.title("🧠 Análisis Inteligente - Florida Tarde")

# ==========================================
# 📂 CONFIGURACIÓN DE ARCHIVOS
# ==========================================
ARCHIVO_ENTRADA = 'david esgrima.xlsx'
HOJA_OBJETIVO = 'TARDE 1 (1)'  # Si no existe, usará la primera hoja automáticamente

# ==========================================
# 🔍 FUNCIÓN DE CARGA ROBUSTA (OpenPyXL)
# ==========================================
@st.cache_data(ttl=3600)
def cargar_datos_desde_excel():
    try:
        wb = load_workbook(ARCHIVO_ENTRADA, data_only=True)
        ws = wb[HOJA_OBJETIVO] if HOJA_OBJETIVO in wb.sheetnames else wb.active
        
        datos = []
        limite_fin = date(2027, 12, 31)  # Permitir fechas hasta 2027
        limite_inicio = date(2010, 1, 1)
        
        # Escanear toda la hoja buscando celdas con relleno ROJO
        for row in ws.iter_rows():
            for cell in row:
                # Verificar color de relleno (rojo estándar o tema)
                fill_rgb = str(getattr(cell.fill.start_color, 'rgb', '') or '').upper()
                is_red = 'FF0000' in fill_rgb or 'C00000' in fill_rgb or fill_rgb.endswith('FF0000')
                
                if is_red:
                    val = str(cell.value or '').strip()
                    # Regex flexible: acepta 10/5/2026, 10-05-26, etc.
                    match = re.search(r'(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{2,4})', val)
                    if match:
                        d, m, a = int(match.group(1)), int(match.group(2)), int(match.group(3))
                        if a < 100: a += 2000
                        try:
                            fecha_dt = datetime(a, m, d)
                            fecha = fecha_dt.date()
                            if limite_inicio <= fecha <= limite_fin:
                                # Buscar el número en la celda de JUSTO ABAJO
                                celda_num = ws.cell(row=cell.row + 1, column=cell.column)
                                num_raw = str(celda_num.value or '').upper().replace('O', '0').replace(' ', '').replace('ERROR:#N/A', '')
                                
                                if num_raw.isdigit() and len(num_raw) >= 2:
                                    fijo = int(num_raw[-2:])  # Tu regla: últimos 2 dígitos
                                    datos.append({
                                        'Fecha': fecha_dt, 
                                        'Numero': fijo, 
                                        'Celda': cell.coordinate
                                    })
                        except: 
                            continue
        
        if not datos:
            return None, "⚠️ No se encontraron celdas ROJAS con fecha válida. Verifica que las fechas tengan relleno rojo y letra blanca."
            
        # Ordenar y eliminar duplicados
        df = pd.DataFrame(datos).drop_duplicates(subset='Fecha').sort_values('Fecha').reset_index(drop=True)
        ultima_fecha = df['Fecha'].max()
        celda_ultima = df[df['Fecha'] == ultima_fecha]['Celda'].iloc[0]
        
        wb.close()
        return df, f"✅ {len(df)} sorteos cargados. Última fecha: `{ultima_fecha.strftime('%d/%m/%Y')}` (Celda: `{celda_ultima}`)"
        
    except Exception as e:
        return None, f"❌ Error leyendo Excel: {e}"

# ==========================================
# 🚀 EJECUCIÓN INICIAL
# ==========================================
df, msg = cargar_datos_desde_excel()
if df is None:
    st.error(msg)
    with st.expander("🔍 Diagnóstico rápido"):
        st.info("1. Verifica que `david esgrima.xlsx` esté en la carpeta raíz del repositorio.\n"
                "2. Asegúrate de que `openpyxl` esté en `requirements.txt`.\n"
                "3. Comprueba que las fechas tengan formato de celda `Relleno: Rojo` y `Fuente: Blanco`.")
    st.stop()
st.success(msg)

# ==========================================
# 📊 CÁLCULO DE MÉTRICAS Y PROYECCIÓN
# ==========================================
fecha_base = df['Fecha'].max()
dia_objetivo = fecha_base + pd.Timedelta(days=1)
nombre_dia = dia_objetivo.strftime('%A')

st.info(f"📅 Fecha base detectada: `{fecha_base.strftime('%d/%m/%Y')}`")
st.info(f"🎯 Próximo sorteo proyectado: **{dia_objetivo.strftime('%d/%m/%Y')} ({nombre_dia})**")

# Columnas base
df['Decena'] = (df['Numero'] // 10).astype(int)
df['Terminacion'] = (df['Numero'] % 10).astype(int)
df['Suma'] = df['Numero'].apply(lambda x: (x//10) + (x%10))
df['DiaSemana'] = df['Fecha'].dt.day_name()

# Cálculo de puntuación para los 100 números
resultados = []
for n in range(0, 100):
    dec, ter = n // 10, n % 10
    
    # 1. Separación (Gap) desde última aparición
    apariciones = df[df['Numero'] == n]['Fecha']
    gap = (fecha_base - apariciones.max()).days if not apariciones.empty else 999
    
    # 2. Tendencia escalonada (últimas 3 decenas)
    ultimas_dec = df.tail(3)['Decena'].tolist()
    tendencia = 0
    if len(ultimas_dec) == 3:
        if ultimas_dec[0] < ultimas_dec[1] < ultimas_dec[2] and dec == ultimas_dec[2] + 1: tendencia = 15
        elif ultimas_dec[0] > ultimas_dec[1] > ultimas_dec[2] and dec == ultimas_dec[2] - 1: tendencia = 15
            
    # 3. Frecuencia histórica en este día de la semana (últimos 6 meses)
    hace_180 = fecha_base - pd.Timedelta(days=180)
    freq_dia = len(df[(df['Fecha'] >= hace_180) & (df['DiaSemana'] == nombre_dia) & (df['Numero'] == n)])
    
    # 4. Fórmula de Puntuación (Ajustable según tu experiencia)
    pts = 0
    if 20 <= gap <= 45: pts += 20      # Zona dulce de separación
    elif gap > 45: pts += 10           # Presión alta
    pts += tendencia                   # Tendencia de decena
    if freq_dia >= 3: pts += 12        # Día fuerte
    if 8 <= (dec + ter) <= 14: pts += 5 # Suma en rango frecuente
    
    resultados.append({
        'Numero': n, 'Decena': dec, 'Terminacion': ter, 
        'Gap': gap, 'Tendencia': tendencia, 'Freq_Dia': freq_dia, 'Puntaje': pts
    })

df_scores = pd.DataFrame(resultados)
df_scores['Estado'] = df_scores['Puntaje'].apply(
    lambda x: '🔴 JUGAR' if x >= 50 else ('🟡 OBSERVAR' if x >= 35 else '⚪ ESPERAR')
)
df_scores = df_scores.sort_values('Puntaje', ascending=False).reset_index(drop=True)

# ==========================================
# 💾 RENDERIZADO EN STREAMLIT
# ==========================================
col1, col2 = st.columns([2, 1])
with col1:
    st.subheader("🏆 Top 10 Proyección")
    st.dataframe(df_scores.head(10), use_container_width=True, hide_index=True)

with col2:
    st.subheader("📊 Resumen del Modelo")
    st.metric("🔴 JUGAR", len(df_scores[df_scores['Estado']=='🔴 JUGAR']))
    st.metric("🟡 OBSERVAR", len(df_scores[df_scores['Estado']=='🟡 OBSERVAR']))
    st.metric("📅 Proyección", f"{dia_objetivo.strftime('%A %d/%m')}")

st.divider()
if st.button("📥 Descargar Excel con análisis completo", use_container_width=True):
    df_scores.to_excel('Prediccion_Inteligente_FL_Tarde.xlsx', index=False)
    st.success("✅ `Prediccion_Inteligente_FL_Tarde.xlsx` generado en la carpeta raíz.")